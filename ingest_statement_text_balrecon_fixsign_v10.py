#!/usr/bin/env python3
"""
Ingest Chase statement text (from `pdftotext -raw`) into the Dashboard workbook.

Fix in this version:
- Correct year assignment for statements that straddle months.
  Uses the end date from the filename (YYYYMMDD).
  Rule:
    • If txn_month == end_month → year = end_year
    • If txn_month == end_month - 1 → year = end_year
    • Special case: if end_month == 1 and txn_month == 12 → year = end_year - 1
"""

import re
import sys
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import hashlib
from datetime import datetime
CALLS = {"parse_dep_add": 0}  # put at module top, once
# Regex patterns
DATE_START_RE = re.compile(r"^\s*(\d{2}/\d{2})")
# ONE capturing group so .findall() returns strings
AMT_RE = re.compile(r"""
    (                                   # capture the whole amount
      -?\s*\$?\s*(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{2}\s*-?   # 1,234.56 or 1234.56, optional $ and trailing -
      | \(\s*\$?\s*(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{2}\s*\) # (1,234.56) style negatives
    )
""", re.X)

SUBTOTAL_RE = re.compile(r"^\s*Total\s+", re.I)
HEADER_RE = re.compile(r"^(DATE|DESCRIPTION|AMOUNT)\b", re.I)
#

# tolerant date at start: MM/DD or M-D (optionally /YYYY) but with real month/day ranges
# Prefer two-digit matches first: 12 -> 12, 31 -> 31, then allow 1..9
DATE_LINE   = re.compile(r'^\s*(1[0-2]|0?[1-9])\s*[/-]\s*(3[01]|[12]\d|0?[1-9])(?:\s*[/-]\s*(\d{2,4}))?')
DATE_TOKEN  = re.compile(r'^\s*(1[0-2]|0?[1-9])\s*[/-]\s*(3[01]|[12]\d|0?[1-9])(?:\s*[/-]\s*(\d{2,4}))?')
DATE_SEARCH = re.compile(r'(1[0-2]|0?[1-9])\s*[/-]\s*(3[01]|[12]\d|0?[1-9])(?:\s*[/-]\s*(\d{2,4}))?')

# Checks header (the real section header)
CHECKS_HEADER = re.compile(r'^\s*CHECKS?\s+PAID\b', re.I)

# Check-line signatures (require a check NUMBER, not just the word "check")
CHECK_NUMBER_INLINE = re.compile(r'\bCHECK\s+#?\d{3,6}\b', re.I)

# Your tolerant check-line patterns from earlier (start-of-line forms)
# (If you already have CHECK_LINE_RE1/CHECK_LINE_RE2, keep those and reuse them here.)
CHECK_LINE_RE1 = re.compile(
    r'^\s*(?:CHECK\s+)?(?P<chkno>\d{3,6})\s*(?:[\^\*]\s*)?(?P<mmdd>\d{1,2}[/-]\d{1,2})\s+(?P<amt>-?\$?\d[\d,]*\.\d{2})',
    re.I
)
CHECK_LINE_RE2 = re.compile(
    r'^\s*(?P<mmdd>\d{1,2}[/-]\d{1,2})\s+(?:CHECK\s+)?(?P<chkno>\d{3,6}).*?(?P<amt>-?\$?\d[\d,]*\.\d{2})',
    re.I
)
#
CHECK_LINE_RE = re.compile(
    r"""^\s*
        (?P<chkno>\d{4,})
        \s+[\^\*]\s+
        (?P<mmdd>\d{2}/\d{2})
        \s+
        (?P<amt>-?\$?\d{1,3}(?:,\d{3})*\.\d{2})
        """,
    re.X
)
# Begin new add
# --- Deposits & Additions (sequential) ---
DATE_START_RE = re.compile(r'^\s*(\d{1,2})\s*[/-]\s*(\d{1,2})')
#
# tolerant MM/DD or M-D, optional /YYYY after the day

DEP_ADD_HDR = re.compile(r'Deposits?\s*(?:&|and)?\s*(?:Other\s+)?(?:Additions?|Credits?)', re.I)
HEADER_ROW = re.compile(r'^\s*DATE\s+DESCRIPTION\s+AMOUNT\b', re.I)

# any obvious “we’re in checks now” signature
CHECK_SNIFF = re.compile(
    r'\bCHECK\b|\b\d{4,6}\b\s*(?:[\^\*]\s*)?\d{1,2}\s*[/-]\s*\d{1,2}', re.I
)

# Matches: Deposits & Additions, Deposits and Additions, Deposits & Other Additions, Deposits & Credits, etc.

# Next-section detectors
NEXT_SEC = re.compile(
    r"^(?:\s*(?:Checks?\s+Paid|Electronic\s+(?:Withdrawals?|Payments?)|"
    r"Debit\s*Card\s*(?:Purchases?|Transactions?)|ATM\s+Withdrawals?))\b",
    re.I
)
MAJOR_STOP = re.compile(
    r"^(?:\s*)(?:CHECKING\s+SUMMARY|CHECK\s+NO\.\s+DESCRIPTION|"
    r"CHECKS?\s+PAID|ATM\s*&?\s*DEBIT\s*CARD\s*WITHDRAWALS?|"
    r"ELECTRONIC\s+WITHDRAWALS?)\b(?!.*\d[\d,]*\.\d{2})",
    re.I
)
ATM_DEBIT_HDR = re.compile(r'ATM\s*&?\s*Debit\s*Card\s*Withdrawals?', re.I)
ELEC_WITH_HDR = re.compile(r'Electronic\s+Withdrawals?', re.I)
# New Balance code
# --- Balance parsing helpers ---
# Prefer the exact phrase you gave us:
CHECKING_SPECIFIC_RE = re.compile(
    r"""
    Chase\s+Better\s+Banking\s+Checking      # literal product+type
    \s+(?P<acct>\d{6,})                      # account number (6+ digits)
    \s+\$?(?P<begin>[\d,]+\.\d{2})           # beginning balance
    \s+\$?(?P<end>[\d,]+\.\d{2})             # ending balance
    """, re.I | re.X
)

# General fallback for any "… Checking <acct> $begin $end" header line
CHECKING_GENERIC_RE = re.compile(
    r"""
    \bChecking
    \s+(?P<acct>\d{6,})
    \s+\$?(?P<begin>[\d,]+\.\d{2})
    \s+\$?(?P<end>[\d,]+\.\d{2})
    \b
    """, re.I | re.X
)

BAL_AMT = r"""
\(?\s*\$?\s*(?:\d{1,3}(?:,\d{3})+|\d+)\.\d{2}\s*\)?   # $1,234.56 or (1,234.56)
"""
BAL_AMT_RE = re.compile(BAL_AMT, re.X)

# Common explicit labels Chase prints on some formats
BEGIN_BAL_RE = re.compile(r"Beginning\s+Balance\s+(" + BAL_AMT + r")", re.I | re.X)
END_BAL_RE   = re.compile(r"(Ending|Closing)\s+Balance\s+(" + BAL_AMT + r")", re.I | re.X)
# Match the checking account header with account number + two amounts
ACCT_LINE_RE = re.compile(
    r"Checking\s+000000714245263\s+\$?([\d,]+\.\d{2})\s+\$?([\d,]+\.\d{2})",
    re.I
)

# Fallback: “... 000000714245263 $2,315.05 $3,244.86” -> last two amounts on the line
# (first = beginning, second = ending)

def _to_float_amt(txt: str) -> float:
    """Parses $1,234.56 or (1,234.56) or -1,234.56 to float with sign."""
    s = txt.strip().replace("$", "").replace(",", "")
    neg = False
    if s.endswith('-'):
        neg, s = True, s[:-1]
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1]
    val = float(s)
    return -val if neg else val

def parse_begin_end_balances(lines: list[str]) -> tuple[float|None, float|None]:
    """
    Returns (begin, end) balances for the CHECKING account on the statement.
    Priority:
      1) Exact literal 'Chase Better Banking Checking … <acct> $begin $end'
      2) Any 'Checking <acct> $begin $end' header line
      3) Labeled 'Beginning Balance …' / 'Ending (or Closing) Balance …'
      4) None, None
    """
    # 1) Exact literal Checking header you provided
    for raw in lines:
        m = CHECKING_SPECIFIC_RE.search(raw)
        if m:
            try:
                b = _to_float_amt(m.group("begin"))
                e = _to_float_amt(m.group("end"))
                return b, e
            except Exception:
                pass  # keep looking if parse fails

    # 2) Generic "Checking <acct> $begin $end" (ignores Savings entirely)
    for raw in lines:
        m = CHECKING_GENERIC_RE.search(raw)
        if m:
            try:
                b = _to_float_amt(m.group("begin"))
                e = _to_float_amt(m.group("end"))
                return b, e
            except Exception:
                pass

    # 3) Labeled lines fallback
    begin_lbl, end_lbl = None, None
    for raw in lines:
        m1 = BEGIN_BAL_RE.search(raw)
        if m1:
            begin_lbl = _to_float_amt(m1.group(1))
        m2 = END_BAL_RE.search(raw)
        if m2:
            end_lbl = _to_float_amt(m2.group(2))
    if begin_lbl is not None and end_lbl is not None:
        return begin_lbl, end_lbl

    # 4) Give up
    return None, None

# --- End RegExes
# --- Statement totals (for reconciliation by side) ---
TOTAL_DEPOSITS_RE = re.compile(
    r"""(?i)
    ^\s*(?:Total\s+)?Deposits?\s*(?:&|and)?\s*(?:Other\s+)?(?:Additions?|Credits?)\s*
    \$?\s*([\d,]+\.\d{2})
    """, re.X
)
TOTAL_WITHDRAWALS_RE = re.compile(
    r"""(?i)
    ^\s*(?:Total\s+)?(?:Withdrawals?|Subtractions?)\s*
    \$?\s*([\d,]+\.\d{2})
    """, re.X
)

def parse_statement_totals(lines: list[str]) -> tuple[float|None, float|None]:
    dep, wd = None, None
    for raw in lines:
        m = TOTAL_DEPOSITS_RE.search(raw)
        if m:
            try: dep = clean_amount(m.group(1))
            except Exception: pass
        m2 = TOTAL_WITHDRAWALS_RE.search(raw)
        if m2:
            try: wd = clean_amount(m2.group(1))
            except Exception: pass
    # Deposits total should be positive; withdrawals positive magnitude
    if dep is not None: dep = abs(dep)
    if wd is not None: wd = abs(wd)
    return dep, wd

# --- Section headers (be permissive; Chase wording varies) ---
SECTION_PATTERNS = {
    "DEPOSITS": re.compile(r"Deposits?\s+(?:&|and)\s+Additions?", re.I),
    "WITHDRAWALS": re.compile(r"(Electronic|ATM|Debit\s*Card)\s+Withdrawals?", re.I),
    "CHECKS": re.compile(r"Checks?\s+Paid", re.I),
    "RETURNS": re.compile(r"(Return?|Returned\s+Items?|Adjustments?)", re.I),
}

DEFAULT_INCOME_KEYS = [
    "ONLINE TRANSFER FROM", "TRANSFER FROM", "DEPOSIT", "PAYROLL",
    "CHECK DEPOSIT", "ATM CHECK DEPOSIT", "INTEREST PAYMENT", "Direct Dep", 
    "REFUND", "REVERSAL", "ACH CREDIT", "RETURN", "Pershing"
]
# Hints to force sign even if income keywords don't match
POS_SIGN_HINTS = [
    "DEPOSIT","ACH CREDIT","CREDIT","PAYROLL","DIRECT DEP",
    "ZELLE CREDIT","MOBILE DEPOSIT","CHECK DEPOSIT","ATM CHECK DEPOSIT",
    "RETURN","RETURNS","REFUND","REVERSAL","ADJUSTMENT","PERSHING",
    "INCOME PMT"  # <= new
]
NEG_SIGN_HINTS = [
    "WITHDRAWAL", "DEBIT", "CARD PURCHASE", "POS PURCHASE",
    "ACH DEBIT", "ATM WITHDRAWAL", "ONLINE PAYMENT", "PAYMENT",
    "TRANSFER TO", "AUTOMATIC PAYMENT"
]
def is_check_txn(line: str) -> bool:
    """True only for actual check transactions or the checks header."""
    if CHECKS_HEADER.match(line):
        return True
    if CHECK_NUMBER_INLINE.search(line):
        return True
    if CHECK_LINE_RE1.match(line) or CHECK_LINE_RE2.match(line):
        return True
    return False

def _norm(s: str) -> str:
    return (s or '').replace('\u00A0',' ').replace('\u2007',' ').replace('\u202F',' ')\
                    .replace('\t',' ').rstrip('\r\n')

def _score_depositish(run_lines: list[str]) -> int:
    dep_kw = [ "DEPOSIT", "CHECK DEPOSIT", "ATM CHECK DEPOSIT", "PAYROLL",
               "DIRECT DEP", "ACH CREDIT", "CREDIT", "ONLINE TRANSFER FROM",
               "TRANSFER FROM", "INTEREST PAYMENT" ]
    neg_kw = [ "CARD PURCHASE", "ATM WITHDRAWAL", "WITHDRAWAL",
               "ONLINE PAYMENT", "PAYMENT", "TRANSFER TO", "DEBIT" ]
    s = 0
    for l in run_lines:
        u = l.upper()
        if any(k in u for k in dep_kw): s += 2
        if any(k in u for k in neg_kw): s -= 2
    return s

def grab_date_run(lines, start_idx, *, debug=False):
    out, started = [], False
    j, N = start_idx + 1, len(lines)
    while j < N:
        raw = _norm(lines[j])
        if debug: print(f"[TRACE] j={j+1} raw={raw!r}")

        # hard stop on next major section header (even pre-run)
        if MAJOR_STOP.match(raw):
            if debug:
                why = "major header after run started" if started else "major header before run started"
                print(f"   ↳ stop: {why}")
            break

        # table header
        if HEADER_ROW.match(raw):
            if debug: print("   ↳ skip header row")
            j += 1; continue

        m = DATE_LINE.match(raw)
        if m:
            if debug: print(f"   ↳ add date-line (mm={m.group(1)}, dd={m.group(2)})")
            out.append(raw); started = True; j += 1; continue

        # only let checks end the run *after* we’ve started
        if is_check_txn(raw):
            if started:
                if debug: print("   ↳ stop: checks header/txn after run started")
                break
            if debug: print("   ↳ pre-run checks summary; continue")
            j += 1; continue

        if started:
            if debug: print("   ↳ stop: non-date after run started")
            break

        if debug: print("   ↳ pre-run noise; continue")
        j += 1

    if debug: print(f"[TRACE] collected {len(out)} date-lines in this run")
    return out
def _count_date_lines(lines_slice: list[str]) -> int:
    c = 0
    for L in lines_slice:
        x = _norm(L)
        if DATE_TOKEN.match(x) or DATE_SEARCH.search(x):
            c += 1
    return c
#---
def find_deposits_window(lines: list[str], *, debug: bool=False) -> tuple[int|None, int|None, list[str]]:
    """
    Find the *detail* Deposits & Additions window, not the summary.
    Strategy: collect all header candidates (1-line and 2-line), build [start,end)
    windows to the next clear section, score each by # of date-looking lines, and
    return the highest-scoring one.
    """
    N = len(lines)
    cands = []

    # 1) single-line headers
    for i in range(N):
        if DEP_ADD_HDR.search(_norm(lines[i])):
            j = i + 1
            while j < N and not NEXT_SEC.search(_norm(lines[j])):
                j += 1
            cands.append((i, j, [_norm(x) for x in lines[i+1:j]]))

    # 2) two-line headers: "Deposits" line, then "Additions/Credits" line
    for i in range(N - 1):
        a = _norm(lines[i]); b = _norm(lines[i+1])
        if re.search(r"\bDeposits?\b", a, re.I) and re.search(r"\b(Additions?|Credits?)\b", b, re.I):
            j = i + 2
            while j < N and not NEXT_SEC.search(_norm(lines[j])):
                j += 1
            cands.append((i, j, [_norm(x) for x in lines[i+2:j]]))

    if not cands:
        return None, None, []

    # pick the window with the most date-looking lines
    best = max(cands, key=lambda t: _count_date_lines(t[2]))
    if debug:
        scored = [(s+1, e, _count_date_lines(win)) for (s, e, win) in cands]
        print("[DEBUG] Deposit candidates (start→end, date-lines):",
              ", ".join(f"{s}->{e} ({k})" for s, e, k in scored))
        print(f"[DEBUG] Chosen deposits window: {best[0]+1} → {best[1]} ({_count_date_lines(best[2])} date-lines)")
    return best[0], best[1], best[2]
# New code
def file_signature(path: Path) -> tuple[str, int, str]:
    b = path.read_bytes()
    return (path.name, len(b), hashlib.sha1(b).hexdigest())

def read_ingest_log(wb):
    name = "Ingest Log"
    cols = ["When","File","Size","SHA1","Parsed","Added","Workbook"]
    if name not in wb.sheetnames:
        ws = wb.create_sheet(name)
        ws.append(cols)
        return set(), ws
    ws = wb[name]
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]: continue
        seen.add((row[1], row[2], row[3]))  # (File, Size, SHA1)
    return seen, ws

def append_ingest_log(ws, dashboard_path: Path, sig, parsed: int, added: int):
    ws.append([
        datetime.now().isoformat(timespec="seconds"),
        sig[0], sig[1], sig[2],
        int(parsed), int(added),
        dashboard_path.name
    ])

# End New code
# Dates at start (tolerant of spaces and '-' slash)

def iter_deposits_additions(lines, *, debug=False):
    """
    Yield each transaction line in 'Deposits & Additions' section.
    - Enter on header (even if header is split across two lines).
    - Skip blanks/subtotals/headers.
    - Stop only on the next clear section header.
    """
    in_sec = False
    header_armed = False   # saw "Deposits", waiting for "Additions" next line

    for idx, raw in enumerate(lines, start=1):
        line = (raw or "").strip()

        if not in_sec:
            # Handle 2-line headers like "DEPOSITS AND" (line break) "ADDITIONS"
            if not header_armed and re.search(r"Deposits?", line, re.I):
                header_armed = True
                continue
            if header_armed and re.search(r"(Additions?|Credits?)", line, re.I):
                in_sec = True
                header_armed = False
                if debug:
                    print(f"[DEBUG] Entered 'Deposits & Additions' at line {idx}: {line!r}")
                continue

            # 1-line header
            if DEP_ADD_HDR.search(line):
                in_sec = True
                if debug:
                    print(f"[DEBUG] Entered 'Deposits & Additions' at line {idx}: {line!r}")
                continue

            continue

        # --- inside the section ---
        if NEXT_SEC.search(line):  # clear next section → exit
            if debug:
                print(f"[DEBUG] Leaving 'Deposits & Additions' at line {idx}: {line!r}")
            break

        # skip noise
        if not line or SUBTOTAL_RE.search(line) or HEADER_RE.search(line):
            continue

        # transaction lines start with date
        if DATE_TOKEN.match(line):
            yield line
        else:
            # ignore non-date lines inside section (don’t break)
            continue

# New parse_dep_add
def parse_dep_add(lines, end_year: int, end_month: int, *, debug: bool=False):
    # --- build candidate runs ---
    runs = []  # <== the list of (header_index, [date-led lines])

    # 1) Header-anchored candidates
    for i, raw in enumerate(lines):
        if DEP_ADD_HDR.search(_norm(raw)):
            r = grab_date_run(lines, i, debug=debug)
            if r:
                runs.append((i, r))

    # 2) Fallback: if no header-based runs, scan for any date-run blocks
    if not runs:
        j, N = 0, len(lines)
        while j < N:
            if DATE_LINE.match(_norm(lines[j])):
                r = grab_date_run(lines, j-1, debug=debug)  # pretend header right before first date
                if r:
                    runs.append((j-1, r))
                k = j + 1
                while k < N and DATE_LINE.match(_norm(lines[k])):
                    k += 1
                j = k
            else:
                j += 1

    if not runs:
        if debug: print("[DEBUG] Deposits: no date-run found")
        return []
    checks_rows = parse_checks_anywhere(lines, end_year, end_month)
    atm_rows    = parse_negative_section(lines, ATM_DEBIT_HDR, end_year, end_month, debug=debug)
    elec_rows   = parse_negative_section(lines, ELEC_WITH_HDR,  end_year, end_month, debug=debug)

    # --- de-dup and filter weird candidates ---
    def _first_date_offset(_lines, start_idx, r):
        if not r: return 10**9
        first = r[0]
        for k in range(start_idx+1, min(start_idx+60, len(_lines))):
            if _norm(_lines[k]) == first:
                return k - start_idx
        return 10**9

    # de-dup by (header index, first-date offset, length)
    seen, uniq = set(), []
    for i, r in runs:
        pos = _first_date_offset(lines, i, r)
        key = (i, pos, len(r))
        if key in seen: 
            continue
        seen.add(key)
        uniq.append((i, r, pos))

    # prefer runs where the first date appears reasonably soon after header
    filtered = [(i, r, pos) for (i, r, pos) in uniq if pos <= 20]
    if not filtered:
        filtered = uniq  # fall back rather than fail

    # pick best: longest run, then smaller first-date offset
    i_best, run_lines, pos = max(filtered, key=lambda t: (_score_depositish(t[1]), len(t[1]), -t[2]))

    if debug:
        print(f"[DEBUG] Deposits chosen run starting near line {i_best+1} with {len(run_lines)} rows (first-date offset {pos})")

    # --- parse the chosen run into rows ---
    rows = []
    for line in run_lines:
        m = DATE_LINE.match(line)
        if not m:
            continue
        mm = int(m.group(1)); dd = int(m.group(2))
        year = assign_year(end_year, end_month, mm)

        tail = line[m.end():].strip()
        amts = AMT_RE.findall(tail)
        if not amts:
            if debug: print(f"   [WARN] no amount on: {line!r}")
            continue
        amt_txt = amts[-1]
        desc = tail[: tail.rfind(amt_txt)].strip()
        amt = clean_amount(amt_txt)

        rows.append((f"{year:04d}/{mm:02d}/{dd:02d}", desc, abs(amt), "DEP_ADD"))
    return rows

def _match_check(line: str):
    m = CHECK_LINE_RE1.match(line) or CHECK_LINE_RE2.match(line) or CHECK_LINE_RE.match(line)
    if not m:
        return None
    # normalize to (chkno, mmdd, amt_text)
    grp = m.groupdict()
    return grp["chkno"], grp["mmdd"], grp["amt"]


def decide_sign(description: str, amt: float):
    """
    Returns (signed_amount, source, match_keyword)
    source ∈ {'pos_hint','neg_hint','income_keywords','fallback_negative'}
    """
    u = (description or "").upper()
# Optional, but nice: directional transfers first
    if re.search(r'\b(TRANSFER|XFER|TRF)\s+FROM\b', u) or ('FROM' in u and 'PERSHING' in u):
        return abs(amt), "transfer_from", "FROM"
    if re.search(r'\b(TRANSFER|XFER|TRF)\s+TO\b', u) or ('TO' in u and 'PERSHING' in u):
        return -abs(amt), "transfer_to", "TO"
    
    for kw in POS_SIGN_HINTS:
        if kw.upper() in u:
            return abs(amt), "pos_hint", kw
    for kw in NEG_SIGN_HINTS:
        if kw.upper() in u:
            return -abs(amt), "neg_hint", kw

    # Fallback to your existing income keyword logic
    for kw in DEFAULT_INCOME_KEYS:
        if kw.upper() in u:
            return abs(amt), "income_keywords", kw

    # Final fallback: treat as expense (negative)
    return -abs(amt), "fallback_negative", None

def clean_amount(a: str) -> float:
    s = a.strip()
    neg = False
    if s.endswith('-'):
        neg = True
        s = s[:-1]
    if s.startswith('(') and s.endswith(')'):
        neg = True
        s = s[1:-1]
    s = s.replace('$', '').replace(',', '').strip()
    if s.startswith('-'):
        neg = True
        s = s.lstrip('-').strip()
    val = float(s)
    return -val if neg else val
def concat_nonempty(dfs, columns=None):
    frames = [d for d in dfs if d is not None and not getattr(d, "empty", True)]
    if not frames:
        return pd.DataFrame(columns=columns)
    return pd.concat(frames, ignore_index=True)


def parse_end_date_from_filename(path: Path):
    """Extract end_year and end_month from filename like 20190107-...pdf/raw."""
    m = re.search(r"(20\d{2})(\d{2})(\d{2})", path.stem)
    if not m:
        return 2018, 12
    y, mth, _ = m.groups()
    return int(y), int(mth)

def assign_year(end_year: int, end_month: int, txn_month: int):
    """
    Assign transaction year based on statement end date.
    Handles normal overlaps (end_month and end_month-1) and December→January year change.
    """
    if txn_month == end_month:
        return end_year
    if txn_month == ((end_month - 1) if end_month > 1 else 12):
        # Prior month of statement
        if end_month == 1 and txn_month == 12:
            return end_year - 1
        return end_year
    # Fallback: assume end_year
    return end_year

def load_rules(csv_path: Path):
    rules = []
    if not csv_path or not csv_path.exists():
        return rules
    df = pd.read_csv(csv_path)
    for _, r in df.iterrows():
        rules.append({
            "keyword": str(r.get("keyword","")),
            "category": str(r.get("category","Other")) or "Other",
            "match_type": str(r.get("match_type","contains")).lower(),
            "case_sensitive": bool(r.get("case_sensitive", False))
        })
    return rules

def categorize_default(description: str) -> str:
    u = (description or "").upper()
    if u.startswith("CHECK #"): return "Checks"
    if "WAL-MART" in u or "WALMART" in u: return "Groceries"
    if u.startswith("TST*"): return "Food & drink"
    if "PERSHING" in u: return "401K transfer"
    if "ONLINE PAYMENT" in u: return "Online payment"
    if "HOME DEPOT" in u or "LOWES" in u: return "Home Repair"
    if "GOLF" in u: return "Golf"
    if "WELLCARE" in u or "CHIROPRACT" in u: return "Health & wellness"
    if "KERA" in u: return "Donations"
    if any(x in u for x in ["KROGER","TRADER JOE","TOM THUMB","CENTRAL MARKET","MARKET STREET","WHOLEFDS","GROC"]): return "Groceries"
    if any(x in u for x in ["QUIKTRIP","SHELL","EXXON","FUEL","GAS"]): return "Gas"
    if any(x in u for x in ["J. JILL","DSW","KOHL","MARSHALLS","REI","ANTHROPOLOGIE","BEDBATH"]): return "Shopping"
    if any(x in u for x in ["DELI","PANERA","FISH AND FIZZ","CAFETERI","SCOTTY P","BREAD ZEPPELIN","MCAF","JASON'S","CAFE"]): return "Food & drink"
    if any(x in u for x in ["NETFLIX","MUSEUM","ARBOR"]): return "Entertainment"
    return "Other"
# --- Deposit subcategory mapping ---
DEPOSIT_SUBCATS = [
    ("Return",        [r"\bREFUND\b", r"\bRETURN\b", r"\bREVERSAL\b", r"\bADJUSTMENT\b"]),  # put first so it wins ties
    ("Payroll",       [r"\bPAYROLL\b", r"\bDIRECT\s+DEP\b", r"\bINCOME\s+PMT\b", r"\bADP\b", r"\bOASISBATCH\b", r"\bNEW\s+YORK\s+LIFE\s+PAYROLL\b"]),
    ("Transfer In",   [r"\bONLINE\s+TRANSFER\s+FROM\b", r"\bTRANSFER\s+FROM\b", r"\bFROM\s+SAV\b", r"\bPERSHING\b"]),
    ("Check Deposit", [r"(?:ATM|MOBILE)?\s*CHECK\s+DEPOSIT"]),
    ("Interest",      [r"\bINTEREST\s+PAYMENT\b"]),
]

def categorize_deposit(desc: str) -> str:
    u = (desc or "").upper()
    for name, patterns in DEPOSIT_SUBCATS:
        for pat in patterns:
            if re.search(pat, u):
                return name
    return "Deposit"  # fallback

def apply_rules(description: str, rules):
    if not rules: return None
    text_cs = description or ""
    text_uc = text_cs.upper()
    for rule in rules:
        kw = rule["keyword"]
        hay = text_cs if rule["case_sensitive"] else text_uc
        needle = kw if rule["case_sensitive"] else kw.upper()
        mt = rule["match_type"]
        try:
            if mt == "contains" and needle in hay: return rule["category"]
            if mt == "startswith" and hay.startswith(needle): return rule["category"]
            if mt == "regex" and re.search(kw, text_cs): return rule["category"]
        except Exception:
            continue
    return None

def infer_sign(description: str, amt: float, income_keys=None) -> float:
    if income_keys is None:
        income_keys = DEFAULT_INCOME_KEYS
    d = (description or "").upper()
    return abs(amt) if any(k in d for k in income_keys) else -abs(amt)

def extract_section_lines(all_lines):
    capturing = False
    out = []
    for raw in all_lines:
        line = (raw or "").strip()

        # Start at the checks total/header
        if not capturing and re.search(r"Total\s+Checks\s+Paid\b", line, re.I):
            capturing = True
            continue

        if not capturing:
            continue

        # Stop when we hit a major section header or a non-check line
        if MAJOR_STOP.match(line) or NEXT_SEC.match(line) or (line and not is_check_txn(line)):
            capturing = False
            continue

        if is_check_txn(line):
            out.append(line)
    return out

# Parse_records
def parse_records_from_lines(lines, end_year: int, end_month: int):
    records = []
    for line in lines:
        if not line or SUBTOTAL_RE.search(line) or HEADER_RE.search(line):
            continue

        # Only accept true check lines here
        if is_check_txn(line):
            m_chk = CHECK_LINE_RE1.match(line) or CHECK_LINE_RE2.match(line) or CHECK_LINE_RE.match(line)
            if not m_chk:
                continue
            chkno = m_chk.group("chkno")
            mmdd  = m_chk.group("mmdd")
            amt   = clean_amount(m_chk.group("amt"))
            mm, dd = map(int, mmdd.split("/"))
            year   = assign_year(end_year, end_month, mm)
            desc   = f"CHECK #{chkno}"
            records.append((f"{year:04d}/{mm:02d}/{dd:02d}", desc, amt))
            continue

        # Not a check → skip (prevents re-parsing deposit/ACH rows)
        continue

    df = pd.DataFrame(records, columns=["Date","Description","Amount"])
    if df.empty: return df
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    return df.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)

def parse_checks_anywhere(all_lines, end_year: int, end_month: int):
    """Scan all lines; extract only true check rows. No section window needed."""
    out = []
    seen = set()  # (chkno, date, amt)
    for raw in all_lines:
        m = _match_check(_norm(raw))
        if not m:
            continue
        chkno, mmdd, amt_txt = m
        mm, dd = [int(x) for x in re.split(r'[/-]', mmdd)]
        year = assign_year(end_year, end_month, mm)
        amt = clean_amount(amt_txt)
        key = (chkno, f"{year:04d}/{mm:02d}/{dd:02d}", round(abs(amt), 2))
        if key in seen:
            continue
        seen.add(key)
        out.append((f"{year:04d}/{mm:02d}/{dd:02d}", f"CHECK #{chkno}", -abs(amt)))
    return out
def parse_negative_section(all_lines, header_re, end_year: int, end_month: int, *, debug=False):
    """Collect date-led rows after header until the next clear section, as negatives."""
    rows = []
    N = len(all_lines)
    for i, raw in enumerate(all_lines):
        if not header_re.search(_norm(raw)): 
            continue
        # walk forward
        j = i + 1
        while j < N:
            line = _norm(all_lines[j])
            if MAJOR_STOP.match(line) or NEXT_SEC.search(line) or HEADER_ROW.match(line) or SUBTOTAL_RE.match(line):
                # stop when we truly enter the next section or hit totals/headers
                if debug: print(f"[TRACE] stop {header_re.pattern!r} at j={j+1}: {line!r}")
                break
            m = DATE_LINE.match(line)
            if m:
                mm, dd = int(m.group(1)), int(m.group(2))
                year = assign_year(end_year, end_month, mm)
                tail = line[m.end():].strip()
                amts = AMT_RE.findall(tail)
                if amts:
                    amt_txt = amts[-1]
                    desc = tail[: tail.rfind(amt_txt)].strip()
                    amt = clean_amount(amt_txt)
                    rows.append((f"{year:04d}/{mm:02d}/{dd:02d}", desc, -abs(amt)))
            j += 1
        # don’t break; some statements repeat the header; we’ll just gather again and dedupe on concat
    return rows
# --- Single-pass stream parser (deposits → checks → atm → electronic) ---
MAJOR_BREAK = re.compile(r'^\s*(TRANSACTION\s+DETAIL|DATE\s+DESCRIPTION\s+AMOUNT\s+BALANCE)\b', re.I)

def parse_stream_simple(lines, end_year: int, end_month: int, *, debug: bool=False):
    """
    Walk the file once. Ignore headers. Only react to:
      • Check rows (explicit check number pattern)
      • Date-led rows (MM/DD…)
    Section order is inferred by appearance: Deposits (first date-run),
    then Checks (any check lines), then ATM (next date-run), then
    Electronic (next date-run). We stop at the Savings 'TRANSACTION DETAIL'
    table to avoid pulling the other account.
    """
    sec_idx = -1         # -1 none, 0 deposits, 1 checks, 2 atm, 3 electronic
    gap_after_atm = False
    out = []             # (Date, Description, Amount, _src)
    N = len(lines)

    def norm(s: str) -> str:
        return (s or '').replace('\u00A0',' ').replace('\u2007',' ').replace('\u202F',' ')\
                        .replace('\t',' ').rstrip('\r\n')

    i = 0
    while i < N:
        raw = lines[i]; line = norm(raw)
        if not line:
            if sec_idx == 2:  # blank gap between ATM and Electronic
                gap_after_atm = True
            i += 1; continue

        # Stop when Savings transaction table begins
        if sec_idx >= 3 and MAJOR_BREAK.match(line):
            if debug: print(f"-> break at {i+1} {line}")
            break

        # 1) Check rows (always negative)
        mchk = CHECK_LINE_RE1.match(line) or CHECK_LINE_RE2.match(line) or CHECK_LINE_RE.match(line)
        if mchk:
            if sec_idx < 1:
                sec_idx = 1
                if debug: print(f"-> enter CHECKS at {i+1} {line}")
            mm, dd = [int(x) for x in re.split(r'[/-]', mchk.group('mmdd'))]
            year   = assign_year(end_year, end_month, mm)
            amt    = clean_amount(mchk.group('amt'))
            desc   = f"CHECK #{mchk.group('chkno')}"
            out.append((f"{year:04d}/{mm:02d}/{dd:02d}", desc, -abs(amt), "CHECKS"))
            i += 1; continue

        # 2) Date-led rows (Deposits/ATM/Electronic)
        mdate = DATE_LINE.match(line)
        if mdate:
            mm, dd = int(mdate.group(1)), int(mdate.group(2))
            year   = assign_year(end_year, end_month, mm)

            # Grab amount token: if the line has a trailing BALANCE amount, take the FIRST amount;
            # otherwise take the LAST amount.
            amatches = list(AMT_RE.finditer(line))
            if not amatches:
                i += 1; continue
            take_first = (' BALANCE' in line.upper()) or (len(amatches) > 1)
            a_span = amatches[0].span(1) if take_first else amatches[-1].span(1)
            amt    = clean_amount(line[a_span[0]:a_span[1]])
            desc   = line[mdate.end():a_span[0]].strip()

            # Section transitions
            if sec_idx == -1:
                sec_idx = 0
                if debug: print(f"-> enter DEPOSITS at {i+1} {line}")
            elif sec_idx == 1:
                sec_idx = 2; gap_after_atm = False
                if debug: print(f"-> enter ATM at {i+1} {line}")
            elif sec_idx == 2 and gap_after_atm:
                sec_idx = 3
                if debug: print(f"-> enter ELECTRONIC at {i+1} {line}")

            # Signing per section
            if sec_idx == 0:
                signed, src = amt, "DEP_ADD"   # keep deposit signs
            elif sec_idx == 2:
                signed, src = -abs(amt), "ATM"
            elif sec_idx >= 3:
                signed, src = -abs(amt), "ELEC"
            else:
                signed, src = amt, "OTHER"

            out.append((f"{year:04d}/{mm:02d}/{dd:02d}", desc, signed, src))
            i += 1; continue

        # Non-date, non-check line inside ATM: mark a gap so the next date-run becomes Electronic.
        if sec_idx == 2:
            gap_after_atm = True

        i += 1

    return out

# End Parse_records

def read_sheet_df(wb, name):
    if name in wb.sheetnames:
        ws = wb[name]
        cols = [c.value for c in ws[1]] if ws.max_row >= 1 else ["Date","Description","Category","Amount"]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)] if ws.max_row >= 2 else []
        df = pd.DataFrame(rows, columns=cols)
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        # Ensure expected columns exist
        for col in ["Description","Category","Amount"]:
            if col not in df.columns:
                df[col] = "" if col != "Amount" else 0.0
        return df[["Date","Description","Category","Amount"]].dropna(subset=["Date"]).reset_index(drop=True)
    return pd.DataFrame(columns=["Date","Description","Category","Amount"])

def write_df(ws, df):
    ws.delete_rows(1, ws.max_row)
    ws.append(["Date","Description","Category","Amount"])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in dataframe_to_rows(df[["Date","Description","Category","Amount"]], index=False, header=False):
        ws.append(r)
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "yyyy/mm/dd"
    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 70
    ws.column_dimensions[get_column_letter(3)].width = 18
    ws.column_dimensions[get_column_letter(4)].width = 12

def merge_dedup(old, new):
    # normalize empties
    old_empty = old is None or getattr(old, "empty", True)
    new_empty = new is None or getattr(new, "empty", True)

    if old_empty and new_empty:
        return pd.DataFrame(columns=["Date","Description","Category","Amount"]).astype({
            "Date": "datetime64[ns]", "Description": "string", "Category": "string", "Amount": "float64",
        })
    comb = new.copy() if old_empty else (old.copy() if new_empty else pd.concat([old, new], ignore_index=True))

    # dtype hygiene
    comb["Date"] = pd.to_datetime(comb["Date"], errors="coerce")
    comb["Description"] = comb["Description"].astype(str)
    if "Category" not in comb.columns:
        comb["Category"] = ""
    comb["Category"] = comb["Category"].astype(str)
    comb["Amount"] = pd.to_numeric(comb["Amount"], errors="coerce")

    # per-sheet dedupe
    comb["dupe_key"] = (
        comb["Date"].dt.strftime("%Y-%m-%d")
        + "||" + comb["Description"]
        + "||" + comb["Amount"].astype(str)
    )
    comb = comb.drop_duplicates(subset=["dupe_key"]).drop(columns=["dupe_key"])
    return comb.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)

# --- Build Monthly & Yearly summaries with safe TOTAL rows ---
def build_summaries_with_totals(stack: pd.DataFrame):
    # Handle empty stack gracefully
    if stack is None or stack.empty:
        monthly_summary = pd.DataFrame(columns=["Month", "Category", "Amount"])
        yearly_summary  = pd.DataFrame(columns=["Year", "Category", "Amount"])
        return monthly_summary, yearly_summary

    s = stack.copy()
    s["Year"]  = s["Date"].dt.year
    s["Month"] = s["Date"].dt.to_period("M").astype(str)

    # Base summaries
    monthly_summary = s.groupby(["Month", "Category"], as_index=False)["Amount"].sum()
    yearly_summary  = s.groupby(["Year",  "Category"], as_index=False)["Amount"].sum()

    # Monthly TOTAL rows
    _month_totals = (
        s.groupby("Month", as_index=False)["Amount"].sum()
         .assign(Category="TOTAL")[["Month", "Category", "Amount"]]
    )
    monthly_summary = pd.concat([monthly_summary, _month_totals], ignore_index=True)
    monthly_summary["_msort"] = pd.PeriodIndex(monthly_summary["Month"], freq="M")
    monthly_summary["_csort"] = (monthly_summary["Category"] == "TOTAL").astype(int)
    monthly_summary = (monthly_summary
                       .sort_values(["_msort", "_csort", "Category"])
                       .drop(columns=["_msort", "_csort"])
                       .reset_index(drop=True))

    # Yearly TOTAL rows
    _year_totals = (
        s.groupby("Year", as_index=False)["Amount"].sum()
         .assign(Category="TOTAL")[["Year", "Category", "Amount"]]
    )
    yearly_summary = pd.concat([yearly_summary, _year_totals], ignore_index=True)
    yearly_summary["_csort"] = (yearly_summary["Category"] == "TOTAL").astype(int)
    yearly_summary = (yearly_summary
                      .sort_values(["Year", "_csort", "Category"])
                      .drop(columns=["_csort"])
                      .reset_index(drop=True))

    return monthly_summary, yearly_summary

def rebuild_sheet(wb, name, df_in):
    if name in wb.sheetnames:
        ws = wb[name]; ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)
    ws.append(list(df_in.columns))
    for c in ws[1]: c.font = Font(bold=True)
    for r in dataframe_to_rows(df_in, index=False, header=False):
        ws.append(r)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Raw text file from `pdftotext -raw`")
    ap.add_argument("--dashboard", required=True, help="Excel dashboard to update")
    ap.add_argument("--rules", default=None, help="Category rules CSV (optional)")
    ap.add_argument("--debug", action="store_true", help="Verbose debug tracing")
    ap.add_argument("--force", action="store_true", help="Re-ingest even if this exact file was logged before")
    args = ap.parse_args()

    input_path     = Path(args.input)
    dashboard_path = Path(args.dashboard)
    rules_path     = Path(args.rules) if args.rules else None

    # Open or create workbook
    wb = load_workbook(dashboard_path) if dashboard_path.exists() else Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and wb.active.max_row <= 1:
        wb.remove(wb.active)

    # Ingest log (initialize first)
    seen, log_ws = read_ingest_log(wb)

    # Duplicate guard
    sig = file_signature(input_path)
    if not args.force and sig in seen:
        print(f"Skip: {sig[0]} already ingested (size={sig[1]}, sha1={sig[2][:10]}…).")
        return

    # Read raw statement text
    lines = input_path.read_text(encoding="utf-8", errors="ignore").splitlines()

    # Parse balances & statement end date
    begin_bal, end_bal = parse_begin_end_balances(lines)
    end_year, end_month = parse_end_date_from_filename(input_path)

    if args.debug:
        # Spot-check: show a small window around Deposits header
        for i, raw in enumerate(lines, 1):
            if re.search(r'deposits?', raw, re.I) or re.search(r'additions?|credits?', raw, re.I):
                print(f"[DEBUG] header-ish line {i}: {raw!r}")
        def _dump_dep_add_window():
            start = None
            for i, raw in enumerate(lines):
                if DEP_ADD_HDR.search(raw) or re.search(r'DEPOSITS?', raw, re.I):
                    start = i; break
            if start is None:
                print("[DEBUG] No Deposits header found at all."); return
            print("[DEBUG] Window from header at line {}:".format(start+1))
            for j in range(start, min(start+40, len(lines))):
                s = lines[j].rstrip("\n")
                print(f"{j+1:04d} {s!r}")
        _dump_dep_add_window()

    # ---------------- PARSE ----------------
    rows = parse_stream_simple(lines, end_year, end_month, debug=args.debug)
    df = pd.DataFrame(rows, columns=["Date","Description","Amount","_src"])
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"]).sort_values("Date").reset_index(drop=True)

    if df.empty:
        print("No transactions parsed.")
        wb.save(dashboard_path)
        append_ingest_log(log_ws, dashboard_path, sig, 0, 0)
        return

    # ---------------- CATEGORIZE ----------------
    rules = load_rules(rules_path) if rules_path else []
    df["Category"] = ""
    is_dep = df["_src"].eq("DEP_ADD")

    # Deposits sub-buckets
    df.loc[is_dep, "Category"] = df.loc[is_dep, "Description"].map(categorize_deposit)

    # Non-deposits via rules/defaults
    mask_rest = ~is_dep
    df.loc[mask_rest, "Category"] = [
        (apply_rules(desc, rules) or categorize_default(desc))
        for desc in df.loc[mask_rest, "Description"].astype(str)
    ]

    # Section defaults if still blank
    df.loc[df["_src"].eq("ATM")    & (df["Category"] == ""), "Category"] = "Card/ATM"
    df.loc[df["_src"].eq("ELEC")   & (df["Category"] == ""), "Category"] = "Electronic"
    df.loc[df["_src"].eq("CHECKS") & (df["Category"] == ""), "Category"] = "Checks"

    # ---------------- SIGNING (section-first) ----------------
    mask_dep = df["_src"].eq("DEP_ADD")
    mask_neg = df["_src"].isin(["CHECKS","ATM","ELEC"])
    df.loc[mask_dep, "Amount"] = df.loc[mask_dep, "Amount"].abs()
    df.loc[mask_neg, "Amount"] = -df.loc[mask_neg, "Amount"].abs()

    # Unknown section rows → conservative keyword rules
    unknown = ~(mask_dep | mask_neg)
    if unknown.any():
        desc = df.loc[unknown, "Description"].astype(str).str.lower().fillna("")
        neg_kw = ["online payment","payment","ach debit","debit card","withdrawal",
                  "atm withdrawal","transfer to","bill pay","zelle to","venmo cashout",
                  "card purchase","pos purchase"]
        pos_kw = ["refund","reversal","return credit","interest","deposit",
                  "zelle from","transfer from","credit","reimburse"]
        if len(desc) > 0:
            neg_mask = desc.str.contains("|".join(map(re.escape, neg_kw)))
            pos_mask = desc.str.contains("|".join(map(re.escape, pos_kw)))
            df.loc[unknown & neg_mask, "Amount"] = -df.loc[unknown & neg_mask, "Amount"].abs()
            df.loc[unknown & pos_mask, "Amount"] = df.loc[unknown & pos_mask, "Amount"].abs()

    # Targeted fix: any 'online payment ... to ... credit card' should be negative
    desc_all = df["Description"].astype(str).str.lower().fillna("")
    fix_mask = desc_all.str.contains(r"online payment.*to .*credit card")
    df.loc[fix_mask, "Amount"] = -df.loc[fix_mask, "Amount"].abs()

    if args.debug:
        amb = df[~df["_src"].isin(["DEP_ADD","CHECKS","ATM","ELEC"])][["Date","Description","Amount","_src"]].head(30)
        if not amb.empty:
            print("\n[DEBUG] Rows without section (signed via keywords):")
            print(amb.to_string(index=False))

    # ---------------- RECON DETAIL (last run) ----------------
    try:
        detail_cols = ["Date","Description","Category","Amount","_src","_sign_source","_sign_keyword"]
        recon_detail = df.copy()
        for col in detail_cols:
            if col not in recon_detail.columns:
                recon_detail[col] = "" if col not in ("Amount","Date") else recon_detail[col]
        rebuild_sheet(wb, "Recon Detail (last run)", recon_detail[detail_cols])
    except Exception as _e:
        if args.debug:
            print(f"[WARN] Recon Detail not written: {_e}")

    # ---------------- MERGE: per-year + All Transactions ----------------
    tx_cols = ["Date","Description","Category","Amount"]
    df_tx = df[tx_cols].copy()

    years_present = sorted({int(y) for y in df_tx["Date"].dt.year.dropna().unique()})
    years_touched = []
    total_added = 0

    for y in years_present:
        yname = str(y)
        ydf_new = df_tx[df_tx["Date"].dt.year == y].copy()
        old_y = read_sheet_df(wb, yname)
        merged_y = merge_dedup(old_y, ydf_new)

        # Write back per-year sheet
        if yname in wb.sheetnames:
            ws_y = wb[yname]
        else:
            ws_y = wb.create_sheet(yname)
        write_df(ws_y, merged_y)

        if len(merged_y) != len(old_y):
            years_touched.append(yname)
            total_added += max(0, len(merged_y) - len(old_y))

    # Update "All Transactions"
    old_all = read_sheet_df(wb, "All Transactions")
    merged_all = merge_dedup(old_all, df_tx)
    if "All Transactions" in wb.sheetnames:
        ws_all = wb["All Transactions"]
    else:
        ws_all = wb.create_sheet("All Transactions")
    write_df(ws_all, merged_all)

    # ---------------- SUMMARIES (full stack) ----------------
    stack = read_sheet_df(wb, "All Transactions")
    monthly_summary, yearly_summary = build_summaries_with_totals(stack)
    rebuild_sheet(wb, "Monthly Summary", monthly_summary)
    rebuild_sheet(wb, "Yearly Summary",  yearly_summary)

    yoy = yearly_summary.pivot_table(
        index="Category", columns="Year", values="Amount",
        aggfunc="sum", fill_value=0
    ).reset_index()
    rebuild_sheet(wb, "YOY Comparison", yoy)

    # ---------------- BALANCE RECONCILIATION ----------------
    try:
        dep_total_calc = float(df.loc[df["Amount"] > 0, "Amount"].sum())
        wd_total_calc  = float((-df.loc[df["Amount"] < 0, "Amount"]).sum())

        stmt_dep, stmt_wd = parse_statement_totals(lines)

        # Statement end date from filename
        m_date = re.search(r"(20\d{2})(\d{2})(\d{2})", input_path.stem)
        stmt_end_ts = pd.NaT
        if m_date:
            y, mth, d = map(int, m_date.groups())
            stmt_end_ts = pd.Timestamp(y, mth, d)

        computed_end = None if begin_bal is None else (begin_bal + dep_total_calc - wd_total_calc)

        sheet_name = "Balance Reconciliation"
        if sheet_name in wb.sheetnames:
            ws_rc = wb[sheet_name]
            cols = [c.value for c in ws_rc[1]] if ws_rc.max_row >= 1 else \
                   ["Statement End","Begin","Deposits (calc)","Withdrawals (calc)","Computed End","Statement End (reported)","Diff",
                    "Stmt Deposits","Stmt Withdrawals","Δ Deposits","Δ Withdrawals"]
            rows = [r for r in ws_rc.iter_rows(min_row=2, values_only=True)] if ws_rc.max_row >= 2 else []
            recon_df = pd.DataFrame(rows, columns=cols)
        else:
            recon_df = pd.DataFrame(columns=["Statement End","Begin","Deposits (calc)","Withdrawals (calc)","Computed End",
                                             "Statement End (reported)","Diff",
                                             "Stmt Deposits","Stmt Withdrawals","Δ Deposits","Δ Withdrawals"])

        new_row = {
            "Statement End": stmt_end_ts,
            "Begin": begin_bal,
            "Deposits (calc)": dep_total_calc,
            "Withdrawals (calc)": wd_total_calc,
            "Computed End": computed_end,
            "Statement End (reported)": end_bal,
            "Stmt Deposits": stmt_dep,
            "Stmt Withdrawals": stmt_wd,
            "Δ Deposits": (None if stmt_dep is None else round(dep_total_calc - stmt_dep, 2)),
            "Δ Withdrawals": (None if stmt_wd is None else round(wd_total_calc - stmt_wd, 2)),
            "Diff": (None if (computed_end is None or end_bal is None) else round(computed_end - end_bal, 2)),
        }

        # Upsert on Statement End
        if "Statement End" in recon_df.columns and not recon_df.empty:
            recon_df = recon_df[recon_df["Statement End"] != new_row["Statement End"]]
        recon_df = pd.concat([recon_df, pd.DataFrame([new_row])], ignore_index=True)

        # Sort by date if possible
        if not recon_df.empty and "Statement End" in recon_df.columns:
            try:
                recon_df["Statement End"] = pd.to_datetime(recon_df["Statement End"], errors="coerce")
                recon_df = recon_df.sort_values("Statement End").reset_index(drop=True)
            except Exception:
                pass

        rebuild_sheet(wb, sheet_name, recon_df)
        if args.debug:
            print(f"[DEBUG] Reconciliation row: begin={begin_bal}, dep={dep_total_calc}, wd={wd_total_calc}, "
                  f"computed_end={computed_end}, end_reported={end_bal}")
    except Exception as _e:
        if args.debug:
            print(f"[WARN] Reconciliation step skipped due to error: {_e}")

    # ---------------- SAVE & LOG ----------------
    wb.save(dashboard_path)
    append_ingest_log(log_ws, dashboard_path, sig, len(df), total_added)

    if years_touched:
        print(f"Done. Updated year sheets: {', '.join(sorted(set(years_touched)))} in {dashboard_path.name}.")
    else:
        print(f"Done. No year sheets changed in {dashboard_path.name}.")
    if args.debug:
        print(f"[DEBUG] begin_bal={begin_bal}  end_bal={end_bal}")

if __name__ == "__main__":
    main()
